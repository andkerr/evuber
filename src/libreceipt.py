import openpyxl

from collections import namedtuple
from datetime import datetime, date, time, timedelta
from email.message import EmailMessage
import email.parser
import email.policy
import imaplib
import itertools
import json
import logging
import mimetypes
import os
from pathlib import Path
from pypdf import PdfReader
import re
import smtplib
import sys
from typing import Iterable
from weasyprint import CSS, HTML


CONFIG_DIR = Path.home() / ".evuber"

UberTrip = namedtuple("UberTrip", ["date", "time", "total", "pdf"])

# global logging config
#
# Client programs can configure the logging behaviour of "library" functions via
# logging.getLogger("libreceipt"). I doubt this is the best strategy, it's just
# the first alternative to adjusting logging.basicConfig, since that affects the
# behaviour of import modules that also use the logging module (for instance,
# setting the loglevel of logging.basicConfig to logging.INFO will exponse logs
# of level logging.INFO and higher produced by _all_ libraries).
#
# TODO: Consider a less kludgey was to let client programs control libreceipt
#       logging.
log = logging.getLogger("libreceipt")
log.addHandler(logging.StreamHandler())


def getconfig(key: str, default: str = None) -> str:
    if (val := os.getenv(key)):
        return val

    config_path = CONFIG_DIR / "config.json"
    if config_path.exists():
        config = json.load(open(config_path, "r"))
        return config.get(key, default)

    return default


def trip_summaries(address: str,
                   password: str,
                   domain: str,
                   since: str,
                   until: str) -> Iterable[tuple[str, str]]:
    descr = "This is not a payment receipt. " \
            "It is a trip summary to acknowledge the completion of the trip. " \
            "You will receive a trip receipt when the payment is processed " \
            "with payment information."
    conds = ['FROM "noreply@uber.com"',
            f'BODY "{descr}"',
             'BODY "Sutton"',
             'BODY "Harvester"',
             'BODY "Burlington, ON"',
             'NOT BODY "Payments"']
    if since and until:
        log.info(f"fetching trip summaries sent to {address} from {since} to {until} (UTC)")
        conds.append(f"SENTSINCE {since}")
        conds.append(f"SENTBEFORE {until}")
    elif since:
        log.info(f"fetching trip summaries sent to {address} since {since} (UTC)")
        conds.append(f"SENTSINCE {since}")
    elif until:
        log.info(f"fetching trip summaries sent to {address} before {until} (UTC)")
        conds.append(f"SENTBEFORE {until}")
    else:
        log.info("fetching all trip summaries sent to {address}")

    with imaplib.IMAP4_SSL(domain) as mail:
        try:
            mail.login(address, password)
        except Exception:
            log.error(email_authfail(address))
            return

        mail.select("INBOX")

        keys = f"({' '.join(conds)})"
        log.info(f"search: {keys}")
        _, data = mail.search(None, keys)
        ids = data[0].split()
        log.info(f"found {len(ids)} matching messages")
        for id in ids:
            _, data = mail.fetch(id, "(RFC822)")
            for section in data:
                if not isinstance(section, tuple):
                    continue
                parser = email.parser.BytesParser(policy=email.policy.default)
                msg = parser.parsebytes(section[1])
                msgid, _ = msg["Message-Id"].strip("<>").split("@")  # <XXXXXXXXXX@domain.com>
                html = msg.get_content().rstrip()
                log.info(f"Message-Id: {msgid} Date: {msg['Date']}")
                yield msgid, html


def parse_trip(receipt: Path) -> UberTrip:
    date_regex = r"(?:January|February|March|April|May|June|July|August|September|October|November|December) \d+, \d+"
    time_regex = r"(\d+/\d+/\d+)?\s*(\d+:\d+\s(?:AM|PM))"
    total_regex = r"Total\s+CA\$(\d+\.\d+)"

    log.info(f"parsing {receipt}")

    text = gettext(receipt)
    pdf = mkpdf(receipt)

    mdate = re.findall(date_regex, text)
    if len(set(mdate)) != 1:
        log.error(f"Expected 1 'Date' field, got {mdate}")
        sys.exit(1)
    date = datetime.strptime(mdate[0], "%B %d, %Y").date()

    mtotal = re.findall(total_regex, text)
    if len(set(mtotal)) != 1:
        log.error(f"Expected 1 'Total' field, got {mtotal}")
    total = float(mtotal[0])

    mtimes = re.findall(time_regex, text)
    time = min(resolve_datetime(date, match) for match in mtimes).time()

    log.info(f"pdf: {pdf}, date: {date}, time: {time}, total: {total}")

    return UberTrip(date, time, total, pdf)


def gettext(receipt: Path) -> str:
    if not receipt.exists():
        return ""

    if receipt.suffix == ".pdf":
        return " ".join(page.extract_text() for page in PdfReader(receipt).pages)
    elif receipt.suffix == ".html":
        with open(receipt, "r") as f:
            html = f.read().rstrip()
            # Replace tags with a space. They can confuse parsing routines, and
            # whitespace ensures that text contained in adjacent tag pairs isn't
            # concatenated by their removal.
            return re.sub(r"<[^>]*>", " ", html)
    else:
        return ""


def mkpdf(receipt: Path) -> Path:
    if not receipt.exists():
        return Path()

    if receipt.suffix == ".pdf":
        return receipt
    elif receipt.suffix == ".html":
        dst = receipt.with_suffix(".pdf")
        doc = HTML(filename=receipt)
        singlepage = CSS(string="@page { size: 8.5in 33in !important; }")
        logging.getLogger("weasyprint").setLevel(100)  # suppress warnings
        doc.write_pdf(dst, stylesheets=[singlepage])
        return dst
    else:
        return Path()


def resolve_datetime(date: date, mtime: tuple) -> datetime:
    d = datetime.strptime(mtime[0], "%m/%d/%y") if mtime[0] else datetime.combine(date, time())
    t = datetime.strptime(mtime[1], "%I:%M %p")
    return d + timedelta(hours=t.hour, minutes=t.minute)


def group_weeks(trips: list[UberTrip]) -> list[list[UberTrip]]:
    def prev_Sunday(t: UberTrip) -> date:
        return t.date - timedelta(days=weekday(t.date))
    trips.sort(key=lambda t: t.date)
    return [(k, list(g)) for k, g in itertools.groupby(trips, prev_Sunday)]


def emit(wkstart: date,
         trips: list[UberTrip],
         template_file: Path,
         with_address: bool) -> None:
    def to_col(wkday):
        sunday_col = "B"
        # The following is _not_ a general solution. Here's hoping the week
        # columns never cross the Z -> AA boundary.
        return chr(ord(sunday_col) + 2 * wkday)

    log.info(f"wkstart: {wkstart}, {len(trips)} trips")

    totals = [0.0 for _ in range(7)]
    for trip in trips:
        totals[weekday(trip.date)] += trip.total

    row = {"date": 10, "customer": 11, "location": 12, "taxi": 27}
    customer = getconfig("NAME")
    location = getconfig("OFFICE")

    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    for i, total in enumerate(totals):
        ws[f"{to_col(i)}{row['date']}"] = date_full(wkstart + timedelta(days=i))
        ws[f"{to_col(i)}{row['customer']}"] = customer
        ws[f"{to_col(i)}{row['location']}"] = location
        if total:
            ws[f"{to_col(i)}{row['taxi']}"] = total

    ws["N2"] = date_full(endofwk(wkstart))
    personal = {
        "NAME": "N3",
        "DEPT": "N4",
        "MANAGER": "N5",
    }

    if with_address:
        personal |= {"ADDRESSLINE1": "N6", "ADDRESSLINE2": "N7"}

    for key, cell in personal.items():
        ws[cell] = getconfig(key, "")

    wkstr = f"{date_short(wkstart)}-{date_short(wkstart + timedelta(days=6))}"
    outdir = Path(f"evuber_{wkstr}")
    outdir.mkdir(exist_ok=True)
    wb.save(outdir / Path(f"{wkstr}.xlsx"))
    for trip in trips:
        outfile = f"{trip.date.strftime('%m%d')}-{trip.time.strftime('%I%p')}.pdf"
        trip.pdf.rename(outdir / outfile)
    return outdir


def date_short(d: date) -> str: return d.strftime("%m%d")
def date_full(d: date) -> str: return d.strftime("%m-%d-%Y")


def weekday(d: date) -> int:
    """Return the day of the week as an integer, where Sunday is 0 and
    Saturday is 6"""
    return d.isoweekday() % 7  # isoweekday() returns an integer from 1 to 7, where
                               # 1 is Monday and 7 is Sunday


def endofwk(d: date) -> date:
    return d + timedelta(days=6)


def senddir(directory: Path,
            fromaddr: str,
            frompass: str,
            toaddr: str,
            domain: str,
            dryrun: bool) -> None:
    log.info(f"Sending {directory} to {toaddr} via {fromaddr}")

    if not directory.name.startswith("evuber_"):
        log.warn(f"It looks like you're trying to send a directory that wasn't created by evuber: {directory}")
    daterange = directory.name[directory.name.find("_"):]
    wkstart, wkend = daterange.split("-")

    msg = EmailMessage()
    msg["Subject"] = f"[Uber Expense Report] {' to '.join([wkstart, wkend])}"
    msg["From"] = fromaddr
    msg["To"] = toaddr

    n = sum(1 for _ in directory.glob("*.pdf"))
    body = f"""\
    I took {n} {'trips' if n > 1 else 'trip'} this week.

    My {'receipts' if n > 1 else 'reciept'} and expense form are attached. If you need anything else, please let me know.
    """
    msg.set_content(body)

    for path in directory.iterdir():
        if not path.is_file():
            continue

        log.info(path.name)
        ctype, encoding = mimetypes.guess_type(path)
        if ctype is None or encoding is not None:
            ctype = "application/octet-stream"
        maintype, subtype = ctype.split("/")
        with open(path, "rb") as f:
            msg.add_attachment(f.read(),
                               maintype=maintype,
                               subtype=subtype,
                               filename=path.name)

    if dryrun:
        return

    with smtplib.SMTP(domain, 587) as s:
        s.starttls()
        try:
            s.login(fromaddr, frompass)
        except Exception:
            log.error(email_authfail(address))
            return
        s.send_message(msg)

def email_authfail(address: str) -> str:
    return f"Failed to login to {address} with the provided credentials. If " \
            "this is your first attempt fetching emails with evuber, you may " \
            "want to confirm whether you need to use an app password " \
            "(http://tinyurl.com/466u6jye)"
